import io
from datetime import datetime
import openpyxl
from fpdf import FPDF
from app.schemas.order import OrderDetail
from app.schemas.dealer import DealerLedgerOut

# Đường dẫn tĩnh tới file Font hỗ trợ tiếng Việt
FONT_PATH = "app/fonts/Roboto-Regular.ttf"

class ExportService:
    
    @staticmethod
    def _format_money(amount: float) -> str:
        return f"{amount:,.0f} đ".replace(",", ".")

    @staticmethod
    def export_orders_to_excel(orders: list[OrderDetail]) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Danh sách Đơn hàng"

        # Headers
        headers = ["Mã đơn", "Ngày tạo", "Trạng thái", "Khách hàng", "Tổng tiền", "Địa chỉ giao"]
        ws.append(headers)

        for order in orders:
            ws.append([
                order.order_no,
                order.created_at.strftime("%d/%m/%Y %H:%M"),
                order.status,
                order.buyer_name or order.customer_id or order.dealer_id or "",
                float(order.total_amount),
                order.shipping_address or ""
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_orders_to_pdf(orders: list[OrderDetail]) -> io.BytesIO:
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        
        try:
            pdf.add_font("Roboto", "", FONT_PATH, uni=True)
            pdf.set_font("Roboto", "", 12)
        except Exception:
            pdf.set_font("Arial", "", 12) # Fallback if font fails

        pdf.cell(200, 10, txt="DANH SÁCH ĐƠN HÀNG", ln=1, align="C")
        pdf.ln(10)

        # Table header
        col_widths = [35, 35, 30, 45, 45]
        headers = ["Mã đơn", "Ngày tạo", "Trạng thái", "Khách hàng", "Tổng tiền"]
        
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 10, str(h), border=1, align="C")
        pdf.ln()

        # Table body
        for order in orders:
            buyer = str(order.buyer_name or "")[:20]
            money_str = ExportService._format_money(float(order.total_amount))
            row = [
                str(order.order_no),
                order.created_at.strftime("%d/%m/%Y"),
                str(order.status),
                buyer,
                money_str
            ]
            for i, val in enumerate(row):
                pdf.cell(col_widths[i], 10, val, border=1, align="C" if i!=3 else "L")
            pdf.ln()

        output = io.BytesIO()
        pdf.output(output) # fpdf2 allows passing BytesIO
        output.seek(0)
        return output

    @staticmethod
    def export_ledger_to_excel(ledgers: list[DealerLedgerOut]) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sổ cái Công nợ"

        headers = ["Thời gian", "Loại giao dịch", "Số tiền", "Nguồn", "Ghi chú"]
        ws.append(headers)

        for l in ledgers:
            ws.append([
                l.created_at.strftime("%d/%m/%Y %H:%M"),
                "GHI NỢ (-)" if l.entry_type == "debit" else "GHI CÓ (+)",
                float(l.amount),
                l.ref_type.upper(),
                l.note or ""
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def export_ledger_to_pdf(ledgers: list[DealerLedgerOut]) -> io.BytesIO:
        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        
        try:
            pdf.add_font("Roboto", "", FONT_PATH, uni=True)
            pdf.set_font("Roboto", "", 12)
        except Exception:
            pdf.set_font("Arial", "", 12)

        pdf.cell(200, 10, txt="SỔ CÁI CÔNG NỢ", ln=1, align="C")
        pdf.ln(10)

        col_widths = [40, 30, 40, 25, 55]
        headers = ["Thời gian", "Loại", "Số tiền", "Nguồn", "Ghi chú"]
        
        for i, h in enumerate(headers):
            pdf.cell(col_widths[i], 10, str(h), border=1, align="C")
        pdf.ln()

        for l in ledgers:
            money_str = ExportService._format_money(float(l.amount))
            type_str = "No (-)" if l.entry_type == "debit" else "Co (+)"
            row = [
                l.created_at.strftime("%d/%m/%Y %H:%M"),
                type_str,
                money_str,
                l.ref_type.upper(),
                str(l.note or "")[:25]
            ]
            for i, val in enumerate(row):
                pdf.cell(col_widths[i], 10, val, border=1, align="C" if i!=4 else "L")
            pdf.ln()

        output = io.BytesIO()
        pdf.output(output)
        output.seek(0)
        return output
